import openpyxl
from openpyxl import Workbook

wb1 = openpyxl.load_workbook("first_excel.xlsx")
wb2 = openpyxl.load_workbook("second_excel.xlsx")
wb3= Workbook()
ws1 = wb1.active
ws2 = wb2.active
ws3 = wb3.active

for i in range(1,3):
    row1_var = ws1.cell(row=1,column=i)
    ws3.cell(row=1,column = i).value = row1_var.value

dict1 = {}
rows1 = ws1.iter_rows(min_row=2,min_col=1)
for i,j in rows1:
    dict1[i.value] = j.value
print(dict1)

dict2 = {}
rows2 = ws2.iter_rows(min_row=2,min_col=1)
for i,j in rows2:
    dict2[i.value] = j.value
print(dict2)
list1=[]
list2=[]
for key,value in dict1.items():
        if key in dict2:
            if dict1[key] == dict2[key]:
                print("True")
                list1.append(key)
                list2.append("True")
            else:
                print("False")
                list1.append(key)
                list2.append("False")
        else:
            list1.append(key)
            list2.append("Bond price not found in file 2")
            print("No value in file 2")

for i in range(len(list1)):
    ws3.cell(row=2+i,column = 1).value = list1[i]

for j in range(len(list2)):
    ws3.cell(row=2+j,column = 2).value = list2[j]

wb3.save("BondPriceResults.xlsx")