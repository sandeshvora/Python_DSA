import openpyxl
wb1 = openpyxl.load_workbook("first_excel.xlsx")
wb2 = openpyxl.load_workbook("second_excel.xlsx")
ws1=wb1['Sheet1']
ws2 = wb2['Sheet1']

# prod_sheet['A7'].value= "Vikram"
# prod_sheet['B7'].value= 75
# print(prod_sheet['B7'].value)

# prod_sheet['C1'].value = "double_score"
# for i in range(2,8):
#     score1 = ws1.cell(row = i, column = 2).value
#     double_score = score1*2
#     ws1.cell(row=i, column = 3).value = double_score

xyz = list(ws1.rows)[1]
for i in xyz:
    print(i.value)

abc = ws1['B']
for i in abc:
    print(i.value)


wb1.save("first_excel.xlsx")
