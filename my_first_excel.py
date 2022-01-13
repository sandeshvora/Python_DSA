import openpyxl
wb1 = openpyxl.load_workbook("first_excel.xlsx")
wb2 = openpyxl.load_workbook("second_excel.xlsx")
# print(wb1.sheetnames)

prod_sheet=wb1['Sheet1']
reg_sheet = wb2['Sheet1']
# print(prod_sheet)
# print(reg_sheet)
# print(prod_sheet['B4'].value)
# print(prod_sheet.cell(row=2,column=1).value)

# wb1.save('bondPriceCompareResults.xlsx')
#
# gen=prod_sheet.iter_cols(min_row=3,max_row=4,min_col=1)
# print(gen)
# for i,v in gen:
#     print(i.value)
# print(gen)
# print(data1)
# for a,b in data1:
#     print(a.value)
rows1 = list(prod_sheet.rows)
print(rows1)
cols1 = list(prod_sheet.columns)
print(cols1)

